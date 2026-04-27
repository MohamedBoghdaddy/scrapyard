"""
Hardened Excel writer for Scrapyard.

Fixes the  lxml.etree.SerialisationError: IO_WRITE  corruption issue by:
  1. Writing to a sibling .tmp file first
  2. Validating the .tmp file with openpyxl before committing
  3. Atomically moving the .tmp file to the final path
  4. Detecting locked output files and falling back to a timestamped name
  5. Retrying up to *max_retries* times before raising / falling back to CSV
  6. Sanitizing all data before writing (illegal XML chars, overflow, etc.)

Public API::

    from utils.excel_writer import safe_excel_write, ExcelWriteResult

    result = safe_excel_write(workbook_dict, Path("output/site"))
    print(result.path, result.format)   # .xlsx or .csv on fallback
"""
from __future__ import annotations

import logging
import os
import shutil
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd

from .data_sanitizer import sanitize_dataframe, sanitize_sheet_name

logger = logging.getLogger(__name__)

# How many rows to spot-check when validating a written workbook
_VALIDATION_SAMPLE_ROWS = 5


@dataclass
class ExcelWriteResult:
    path: Path
    format: str          # "excel" or "csv"
    fallback_used: bool  # True when CSV fallback was triggered
    rows_written: int
    sheets: List[str]


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def safe_excel_write(
    workbook: Dict[str, pd.DataFrame],
    output_path: Path,
    *,
    max_retries: int = 3,
    fallback_to_csv: bool = True,
    no_excel_fallback: bool = False,
) -> ExcelWriteResult:
    """
    Write *workbook* (sheet_name → DataFrame) to *output_path*.xlsx safely.

    Parameters
    ----------
    workbook:
        Ordered dict of sheet names → DataFrames.
    output_path:
        Target path **without** extension.  The writer adds .xlsx (or .csv on
        fallback).
    max_retries:
        How many write attempts to make before giving up / falling back.
    fallback_to_csv:
        If True (default), fall back to CSV on Excel failure.
    no_excel_fallback:
        If True, raise on Excel failure instead of falling back.
    """
    xlsx_path = output_path.with_suffix(".xlsx")
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    total_rows = sum(len(df) for df in workbook.values())
    last_exc: Optional[Exception] = None

    for attempt in range(1, max_retries + 1):
        tmp_path: Optional[Path] = None
        try:
            tmp_path = _write_to_temp(workbook, xlsx_path)
            _validate_xlsx(tmp_path)
            final_path = _atomic_move(tmp_path, xlsx_path)
            logger.info(
                "Excel saved → %s  (%d rows, %d sheets, attempt %d)",
                final_path, total_rows, len(workbook), attempt,
            )
            return ExcelWriteResult(
                path=final_path,
                format="excel",
                fallback_used=False,
                rows_written=total_rows,
                sheets=list(workbook.keys()),
            )

        except Exception as exc:
            last_exc = exc
            logger.warning(
                "Excel write attempt %d/%d failed: %s",
                attempt, max_retries, exc,
            )
            _cleanup_temp(tmp_path)
            if attempt < max_retries:
                time.sleep(0.5 * attempt)   # brief back-off between retries

    # All retries exhausted
    logger.error("Excel export failed after %d attempts: %s", max_retries, last_exc)

    if no_excel_fallback:
        raise RuntimeError(
            f"Excel export failed after {max_retries} attempts"
        ) from last_exc

    if not fallback_to_csv:
        raise RuntimeError(
            f"Excel export failed and fallback is disabled"
        ) from last_exc

    return _csv_fallback(workbook, output_path, total_rows)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_to_temp(
    workbook: Dict[str, pd.DataFrame],
    xlsx_path: Path,
) -> Path:
    """Write sanitized workbook to a temp file in the same directory."""
    # Use a .xlsx suffix for the temp file so openpyxl accepts it
    fd, tmp_str = tempfile.mkstemp(suffix=".tmp.xlsx", dir=xlsx_path.parent)
    os.close(fd)
    tmp_path = Path(tmp_str)

    with pd.ExcelWriter(str(tmp_path), engine="openpyxl") as writer:
        for raw_name, df in workbook.items():
            sheet_name = sanitize_sheet_name(raw_name)
            clean_df = sanitize_dataframe(df)
            clean_df.to_excel(writer, index=False, sheet_name=sheet_name)
            _autosize_cols(writer.sheets[sheet_name])

    return tmp_path


def _validate_xlsx(tmp_path: Path) -> None:
    """Open the temp file with openpyxl to confirm it is not corrupt."""
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(str(tmp_path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    # Spot-check: read a few cells from the first sheet
    if sheet_names:
        ws = wb[sheet_names[0]]
        rows_read = 0
        for row in ws.iter_rows(max_row=_VALIDATION_SAMPLE_ROWS):
            _ = [cell.value for cell in row]
            rows_read += 1
    wb.close()
    logger.debug("Excel validation OK: %d sheets", len(sheet_names))


def _atomic_move(tmp_path: Path, target: Path) -> Path:
    """
    Move tmp → target atomically.  If target is locked (PermissionError on
    Windows), fall back to a timestamped filename in the same directory.
    """
    try:
        shutil.move(str(tmp_path), str(target))
        return target
    except PermissionError:
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        alt = target.parent / f"{target.stem}_{ts}.xlsx"
        shutil.move(str(tmp_path), str(alt))
        logger.warning(
            "Output file locked (%s); saved to alternate path: %s",
            target, alt,
        )
        return alt


def _cleanup_temp(tmp_path: Optional[Path]) -> None:
    """Remove temp file if it exists. Matches both .xlsx.tmp and .tmp.xlsx patterns."""
    if tmp_path and tmp_path.exists():
        try:
            tmp_path.unlink()
        except Exception as exc:
            logger.debug("Could not remove temp file %s: %s", tmp_path, exc)


def _find_and_cleanup_temps(directory: Path) -> None:
    """Remove any leftover temp Excel files in *directory*."""
    for f in directory.glob("*.tmp.xlsx"):
        _cleanup_temp(f)


def _csv_fallback(
    workbook: Dict[str, pd.DataFrame],
    output_path: Path,
    total_rows: int,
) -> ExcelWriteResult:
    """Write the 'products' sheet (or first sheet) as CSV fallback."""
    logger.warning("Excel export failed → falling back to CSV")
    csv_path = output_path.with_suffix(".csv")

    # Prefer the products sheet; fall back to the first available
    # NOTE: cannot use `or` — DataFrame truth value is ambiguous
    df = workbook.get("products")
    if df is None:
        df = next(iter(workbook.values()))
    clean = sanitize_dataframe(df)
    clean.to_csv(str(csv_path), index=False, encoding="utf-8-sig")

    logger.info("CSV fallback written → %s", csv_path)
    return ExcelWriteResult(
        path=csv_path,
        format="csv",
        fallback_used=True,
        rows_written=len(df),
        sheets=["products"],
    )


def _autosize_cols(worksheet) -> None:
    for col in worksheet.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=10,
        )
        worksheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


# ---------------------------------------------------------------------------
# Large-dataset splitting
# ---------------------------------------------------------------------------

def split_and_write_excel(
    workbook: Dict[str, pd.DataFrame],
    output_path: Path,
    *,
    max_rows_per_file: int = 20_000,
    max_retries: int = 3,
    fallback_to_csv: bool = True,
    no_excel_fallback: bool = False,
) -> List[ExcelWriteResult]:
    """
    If the 'products' sheet exceeds *max_rows_per_file* rows, write multiple
    part files.  All other sheets (metadata, nlp, etc.) are included in each
    part.

    Returns a list of ExcelWriteResult, one per part.
    """
    products_df = workbook.get("products", pd.DataFrame())
    total = len(products_df)

    if total <= max_rows_per_file:
        return [
            safe_excel_write(
                workbook,
                output_path,
                max_retries=max_retries,
                fallback_to_csv=fallback_to_csv,
                no_excel_fallback=no_excel_fallback,
            )
        ]

    # Split into parts
    n_parts = math.ceil(total / max_rows_per_file)
    logger.info(
        "Large dataset: %d rows → splitting into %d parts of ≤%d rows",
        total, n_parts, max_rows_per_file,
    )

    results: List[ExcelWriteResult] = []
    for part_idx in range(n_parts):
        start = part_idx * max_rows_per_file
        end   = min(start + max_rows_per_file, total)
        part_products = products_df.iloc[start:end].reset_index(drop=True)

        # Also slice nlp_enrichment if present
        part_workbook: Dict[str, pd.DataFrame] = {}
        for sheet, df in workbook.items():
            if sheet == "products":
                part_workbook[sheet] = part_products
            elif sheet == "nlp_enrichment" and len(df) == total:
                part_workbook[sheet] = df.iloc[start:end].reset_index(drop=True)
            else:
                part_workbook[sheet] = df  # metadata/vendors shared across parts

        part_path = output_path.parent / f"{output_path.name}_part_{part_idx + 1}"
        result = safe_excel_write(
            part_workbook,
            part_path,
            max_retries=max_retries,
            fallback_to_csv=fallback_to_csv,
            no_excel_fallback=no_excel_fallback,
        )
        logger.info("Part %d/%d → %s", part_idx + 1, n_parts, result.path)
        results.append(result)

    return results


# Make math available without a top-level import (avoids circular issues)
import math  # noqa: E402 – intentional late import
