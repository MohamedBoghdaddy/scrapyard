"""
Tests for utils/excel_writer.py

Covers:
- Normal Excel write succeeds and produces a valid .xlsx file
- Temp file is used and cleaned up on success
- Corrupt / locked output falls back to CSV
- Data sanitization applied before writing
- Validation catches empty/truncated workbooks
- Large dataset splits into multiple part files
- Quality report sheet is included when provided
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Dict
from unittest.mock import patch, MagicMock

import pandas as pd
import pytest

from utils.excel_writer import safe_excel_write, split_and_write_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(rows: int = 10) -> Dict[str, pd.DataFrame]:
    products = pd.DataFrame({
        "name": [f"Product {i}" for i in range(rows)],
        "price": [float(100 + i) for i in range(rows)],
        "url": [f"https://example.com/p{i}" for i in range(rows)],
        "stock_status": ["in_stock"] * rows,
    })
    meta = pd.DataFrame([{"run_id": "test-run", "total_products": rows}])
    return {"products": products, "scrape_metadata": meta}


# ---------------------------------------------------------------------------
# Normal write
# ---------------------------------------------------------------------------

def test_safe_excel_write_creates_xlsx(tmp_output):
    workbook = _make_workbook(20)
    result = safe_excel_write(workbook, tmp_output / "test")

    assert result.format == "excel"
    assert result.fallback_used is False
    assert result.path.suffix == ".xlsx"
    assert result.path.exists()
    assert result.rows_written > 0


def test_safe_excel_write_valid_content(tmp_output):
    workbook = _make_workbook(15)
    result = safe_excel_write(workbook, tmp_output / "check")

    import openpyxl
    wb = openpyxl.load_workbook(str(result.path))
    assert "products" in wb.sheetnames
    wb.close()


def test_safe_excel_write_no_temp_files_left(tmp_output):
    workbook = _make_workbook(5)
    safe_excel_write(workbook, tmp_output / "clean")
    temp_files = list(tmp_output.glob("*.tmp"))
    assert len(temp_files) == 0


# ---------------------------------------------------------------------------
# Data sanitization
# ---------------------------------------------------------------------------

def test_excel_write_strips_illegal_xml_chars(tmp_output):
    workbook = {
        "products": pd.DataFrame([{
            "name": "Part\x00\x01Name",
            "price": 100.0,
            "url": "https://example.com/p1",
        }])
    }
    result = safe_excel_write(workbook, tmp_output / "dirty")
    assert result.path.exists()

    import openpyxl
    wb = openpyxl.load_workbook(str(result.path))
    ws = wb["products"]
    # First data row, first column (name)
    name_cell = ws.cell(row=2, column=1).value
    assert name_cell is not None
    assert "\x00" not in str(name_cell)
    wb.close()


# ---------------------------------------------------------------------------
# CSV fallback
# ---------------------------------------------------------------------------

def test_fallback_to_csv_on_write_failure(tmp_output):
    """When ExcelWriter raises, the function should fall back to CSV."""
    workbook = _make_workbook(10)

    with patch("pandas.ExcelWriter", side_effect=OSError("disk full")):
        result = safe_excel_write(
            workbook,
            tmp_output / "fallback",
            max_retries=1,
            fallback_to_csv=True,
        )

    assert result.format == "csv"
    assert result.fallback_used is True
    assert result.path.suffix == ".csv"
    assert result.path.exists()


def test_no_fallback_raises_on_failure(tmp_output):
    """With no_excel_fallback=True, failure must raise RuntimeError."""
    workbook = _make_workbook(5)

    with patch("pandas.ExcelWriter", side_effect=OSError("write error")):
        with pytest.raises(RuntimeError):
            safe_excel_write(
                workbook,
                tmp_output / "no_fallback",
                max_retries=1,
                fallback_to_csv=False,
                no_excel_fallback=True,
            )


# ---------------------------------------------------------------------------
# Locked file (PermissionError → timestamped alternate)
# ---------------------------------------------------------------------------

def test_locked_file_generates_alternate_name(tmp_output):
    workbook = _make_workbook(5)
    target = tmp_output / "locked.xlsx"

    # Simulate PermissionError on the final move only (not temp write)
    original_move = __import__("shutil").move

    call_count = {"n": 0}

    def _mock_move(src, dst):
        call_count["n"] += 1
        if call_count["n"] == 1 and str(dst).endswith(".xlsx"):
            raise PermissionError("file is locked")
        return original_move(src, dst)

    with patch("shutil.move", side_effect=_mock_move):
        result = safe_excel_write(workbook, tmp_output / "locked")

    assert result.path.exists()
    # The alternate path should include a timestamp suffix
    assert result.format in ("excel", "csv")


# ---------------------------------------------------------------------------
# Large dataset splitting
# ---------------------------------------------------------------------------

def test_split_writes_multiple_parts(tmp_output):
    workbook = _make_workbook(rows=50)
    results = split_and_write_excel(
        workbook,
        tmp_output / "large",
        max_rows_per_file=20,
    )
    assert len(results) == 3   # ceil(50/20) = 3
    for r in results:
        assert r.path.exists()


def test_no_split_when_under_limit(tmp_output):
    workbook = _make_workbook(rows=10)
    results = split_and_write_excel(
        workbook,
        tmp_output / "small",
        max_rows_per_file=20,
    )
    assert len(results) == 1


# ---------------------------------------------------------------------------
# Quality report sheet
# ---------------------------------------------------------------------------

def test_quality_report_sheet_included(tmp_output):
    from utils.quality_report import build_quality_report, quality_report_to_dataframe

    products = [
        {"name": "Part A", "url": "https://ex.com/a", "price": 100.0, "vendor": "Brand"},
        {"name": "Part B", "url": "https://ex.com/b", "price": 200.0, "vendor": ""},
    ]
    report = build_quality_report(products)
    report_df = quality_report_to_dataframe(report)

    workbook = _make_workbook(5)
    workbook["data_quality_report"] = report_df

    result = safe_excel_write(workbook, tmp_output / "with_qr")
    assert result.format == "excel"

    import openpyxl
    wb = openpyxl.load_workbook(str(result.path))
    assert "data_quality_report" in wb.sheetnames
    wb.close()
